import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import lmfit
from lmfit import minimize, Minimizer, Parameters, Parameter, report_fit
import openpyxl
import files, fitting, plotting, excel
import Tkinter,tkFileDialog
import time


def process(row_count, fileiter, datafile, ta, file_list):
    print(datafile)
    Analyzed_counter=0
    data = pd.read_csv(datafile, sep='\t', header=[0, 1])
    wbname = datafile.rsplit('\\', 3)
    file = "../PLM-5mMtet-"+wbname[-2][:-5].rsplit('_', 2)[0]+".xlsx"
    
    try:
        wb = openpyxl.load_workbook(file)
        Sheets = wb.sheetnames
        print(10)
        if wbname[-2][:-5].rsplit('_', 2)[1] not in Sheets:
            excel.new_sheet(wb, wbname[-2][:-5].rsplit('_', 2)[1], ta, row_count)
            
    except IOError:
        file = "../Trial.xlsx"
        wb = openpyxl.load_workbook(file)
        Sheets = wb.sheetnames
        excel.new_sheet(wb, wbname[-2][:-5].rsplit('_', 2)[1], ta, row_count)
        print(20)        
    active_sheet = wb[wbname[-2][:-5].rsplit('_', 2)[1]]
        
    Analyzed_counter, fileiter = excel.analyzed(active_sheet, wbname[-1], fileiter, Analyzed_counter, row_count)

    if Analyzed_counter == 0:
        # Defining the plot
        fig, ax, ax2, ax3 = plotting.make_plot()

        # Cleaning the data into a readable format
        Gt, corrt, G, corr, fileiter, error = fitting.clean_data(data, fileiter)
        if error == 1:
            return fileiter
        
        params = fitting.generate_parameters(ta, Gt)

        # Fitting
        fit = lmfit.minimize(fitting.residual, params, args=(corr, G, ta))

        # Final fit model
        for i in enumerate(corr):
            if i[0] < 1:
                modelGt = fitting.modeld(fit.params, i[1], ta)                
            else:
                modelGt = np.append(modelGt,
                                    fitting.modeld(fit.params, i[1], ta))
        #print(modelGt)
        # Saving model paramters
        a = []
        for i in range(0, np.size(ta)):
            a = np.append(a, fit.params['a'+str(i)])
            #print(fit.params['a'+str(i)].value)

        # Exception if error too large
        if np.mean(G-modelGt) > 1:
            print('error'+wbname[-1])
        #print(G)
        #Plot final graph
        plotting.add_data(ta, a, corrt, G, Gt, corr, modelGt, ax , ax2, ax3, datafile, fig)


        Nmol = float(fit.params['N'])
        '''Td = float(fit.params['td'])
        NA = float(fit.params['Na'])
        Nmol2 = float(fit.params['N2'])
        Td2 = float(fit.params['td2'])
        NA2 = float(fit.params['Na2'])
        Tt = float(fit.params['tT'])
        T = float(fit.params['T'])
        '''

        # Export data in excel
        excel.export(active_sheet, np.append(a,fit.params['N1'].value), wbname[-1], row_count, Nmol, ta)

        fileiter += 1
        wb.save("../PLM-5mMtet-"+wbname[-2][:-5].rsplit('_', 2)[0]+".xlsx")
        print('Finis'+"  "+wbname[-1]+" "+str(len(file_list)-fileiter)+' left')

        
    
    return fileiter

def main():
    directory = "E:\Vidur\Analysis ongoing\FCS-arranged\Non RNAi"
    root = Tkinter.Tk()
    folder = tkFileDialog.askdirectory(initialdir = directory,parent=root,title='Choose a folder')
    root.withdraw()
    
    row_count = 1
    fileiter = 0
    file_list = files.get_files(folder, only=[".dat", "100", "PLM"], skip=["codes",".png","away"])
    ta = np.logspace(-2, 1.5, num=40)
    start_time_final = time.time()
    
    for datafile in file_list:
        start_time = time.time()
        fileiter = process(row_count, fileiter, datafile, ta, file_list)
        print("--- %s seconds ---" % (time.time() - start_time))
    print("--- %s seconds to complete---" % (time.time() - start_time_final))
if __name__ == '__main__':
    main()
